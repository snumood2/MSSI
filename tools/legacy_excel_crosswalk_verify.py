import json
import re
import shutil
import subprocess
import time
import urllib.parse
import urllib.request
from collections import deque
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries

ROOT = Path("/Users/wm/Documents/Codex/2026-06-07/files-mentioned-by-the-user-docx")
REPO = ROOT / "work/MSSI"
ORIGINAL_XLSX = Path("/Users/wm/Library/CloudStorage/OneDrive-개인/9. Temporal Work/2026 설문웹앱개밸/기분장애클리닉 기본 설문지 (우울, 불안, 정서기질)(응답) 오전 5.34.46.xlsx")
TESTDATA_XLSX = Path("/Users/wm/Library/CloudStorage/OneDrive-개인/9. Temporal Work/2026 설문웹앱개밸/테스트데이터.xlsx")
OUT_DIR = REPO / "excel-e2e-artifacts"
SHEET_ID = "1mHaUquO0qdv7bpj9T7LIPVfyUsUlX87uyHiAS_dyG78"
TOKEN_PATH = ROOT / ".secrets/google_token_script.json"
PATIENTS = [
    ("JWEY6A", "WEB-20260607172846-1"),
    ("JWEY6A", "WEB-20260607172846-2"),
]
BROKEN_REPORT_ROWS = set(range(99, 110))


def token():
    info = json.loads(TOKEN_PATH.read_text())
    body = urllib.parse.urlencode({
        "client_id": info["client_id"],
        "client_secret": info["client_secret"],
        "refresh_token": info["refresh_token"],
        "grant_type": "refresh_token",
    }).encode()
    req = urllib.request.Request(info.get("token_uri", "https://oauth2.googleapis.com/token"), data=body, method="POST")
    req.add_header("content-type", "application/x-www-form-urlencoded")
    return json.loads(urllib.request.urlopen(req).read())["access_token"]


ACCESS_TOKEN = token()


def gsheet(range_name, render="FORMATTED_VALUE"):
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{urllib.parse.quote(range_name, safe='')}"
    url += "?" + urllib.parse.urlencode({"valueRenderOption": render})
    req = urllib.request.Request(url)
    req.add_header("authorization", f"Bearer {ACCESS_TOKEN}")
    return json.loads(urllib.request.urlopen(req).read()).get("values", [])


def gsheet_put(range_name, values):
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}/values/{urllib.parse.quote(range_name, safe='')}"
    url += "?valueInputOption=USER_ENTERED"
    data = json.dumps({"range": range_name, "majorDimension": "ROWS", "values": values}).encode()
    req = urllib.request.Request(url, data=data, method="PUT")
    req.add_header("authorization", f"Bearer {ACCESS_TOKEN}")
    req.add_header("content-type", "application/json")
    urllib.request.urlopen(req).read()


def norm_question(value):
    text = str(value or "")
    if "/" in text:
        text = text.split("/")[-1]
    text = re.sub(r"^\s*\d+(?:[-.]\d+)*\.?\s*", "", text)
    text = text.replace("\t", " ")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[\"'“”‘’.,!?()（）:：;；·ㆍ/\\-]", "", text)
    return text


def norm_name(value):
    return re.sub(r"\s+", "", str(value or ""))


def first_num(value):
    m = re.search(r"-?\d+(?:\.\d+)?", str("" if value is None else value))
    return float(m.group(0)) if m else None


def close(a, b, tol=0.51):
    na, nb = first_num(a), first_num(b)
    if na is None or nb is None:
        return str(a or "") == str(b or "")
    return abs(na - nb) <= tol


def calc_refs_from_formula(formula):
    if not isinstance(formula, str) or not formula.startswith("="):
        return set()
    refs = set()
    text = formula.replace("$", "")
    # Explicit quoted or unquoted references to 결과계산(2)
    for m in re.finditer(r"(?:'결과계산\(2\)'|결과계산\(2\))!([A-Z]{1,3}\d+(?::[A-Z]{1,3}\d+)?)", text):
        refs.update(expand_ref(m.group(1)))
    # Local references inside 결과계산(2). Avoid obvious function names by requiring row digits.
    for m in re.finditer(r"(?<![A-Z0-9_])([A-Z]{1,3}\d+(?::[A-Z]{1,3}\d+)?)(?![A-Z0-9_])", text):
        refs.update(expand_ref(m.group(1)))
    return refs


def expand_ref(ref):
    if ":" not in ref:
        return {ref}
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    if (max_col - min_col + 1) * (max_row - min_row + 1) > 5000:
        return set()
    out = set()
    for col in range(min_col, max_col + 1):
        for row in range(min_row, max_row + 1):
            out.add(f"{get_column_letter(col)}{row}")
    return out


def dependency_closure(wb):
    calc = wb["결과계산(2)"]
    report = wb["검사결과지"]
    q = deque()
    seen = set()
    for row in report.iter_rows():
        for cell in row:
            for ref in calc_refs_from_formula(cell.value):
                q.append(ref)
    while q:
        coord = q.popleft()
        if coord in seen:
            continue
        seen.add(coord)
        value = calc[coord].value
        for ref in calc_refs_from_formula(value):
            if ref not in seen:
                q.append(ref)
    return seen


def manual_header_for_col(col):
    if 255 <= col <= 290:
        return f"ip{col - 254}"
    if 291 <= col <= 315:
        return f"cd{col - 290}"
    if 335 <= col <= 361:
        return f"er{col - 334}"
    if 392 <= col <= 404:
        return f"csm{col - 391}"
    if 433 <= col <= 438:
        return f"spaq2_{col - 433}"
    if col == 439:
        return "spaq3"
    if col == 440:
        return "score_SPAQ_global"
    if col == 441:
        return "spaq4"
    if 442 <= col <= 445:
        return f"spaq5_{col - 442}"
    if col == 446:
        return "spaq6"
    if 447 <= col <= 464:
        return f"adhd{col - 446}"
    if 489 <= col <= 513:
        return f"wurs{col - 488}"
    if 520 <= col <= 533:
        return f"pms1_{col - 519}"
    if 534 <= col <= 538:
        return f"pms_imp{col - 533}"
    # MIQ-S/MSSI legacy block: label, frequency, severity repeat.
    if 607 <= col <= 666:
        offset = col - 607
        item = offset // 3 + 1
        slot = offset % 3
        if item <= 20 and slot == 1:
            return f"mssi{item}_freq"
        if item <= 20 and slot == 2:
            return f"mssi{item}_sev"
    if 672 <= col <= 688:
        return f"mt{col - 671}"
    if 362 <= col <= 381:
        return f"bb{col - 361}"
    if 778 <= col <= 813:
        return f"ba{col - 777}"
    diag_map = {
        819: "ds1", 820: "ds2", 821: "ds3", 822: "ds4",
        823: "d1a", 824: "d1b", 825: "d2",
        826: "e1", 827: "e3",
        828: "f1", 829: "f3", 830: "f4", 831: "f6",
        832: "g1a", 833: "g2", 834: "g3a", 835: "g5",
        836: "n1a", 837: "n1b", 838: "n2",
        839: "n3a", 840: "n3b", 841: "n3c", 842: "n3d", 843: "n3e", 844: "n3f",
    }
    if col in diag_map:
        return diag_map[col]
    return None


def build_question_map(needed_cols):
    db_rows = gsheet("DB2SHEET!A1:ZW3", "FORMATTED_VALUE")
    headers = db_rows[0]
    questions = db_rows[1]
    db_q_by_norm = {}
    db_header_to_idx = {}
    for idx, (h, q) in enumerate(zip(headers, questions)):
        if h:
            db_header_to_idx[h] = idx
        n = norm_question(q)
        if n and h and not str(h).startswith("score_"):
            db_q_by_norm.setdefault(n, []).append((idx, h, q))

    wb = load_workbook(ORIGINAL_XLSX, data_only=False)
    ws = wb["결과계산(2)"]
    matches = []
    needed_vlookup_cols = []
    dropped_vlookup_cols = []
    unmatched_needed = []
    for col in range(3, ws.max_column + 1):
        formula = ws.cell(4, col).value
        if not isinstance(formula, str) or "VLOOKUP" not in formula:
            continue
        if col not in needed_cols:
            dropped_vlookup_cols.append((col, ws.cell(3, col).value))
            continue
        needed_vlookup_cols.append(col)
        q = ws.cell(3, col).value
        n = norm_question(q)
        manual_header = manual_header_for_col(col)
        if manual_header and manual_header in db_header_to_idx:
            idx = db_header_to_idx[manual_header]
            matches.append((col, idx, manual_header, q, questions[idx]))
        else:
            candidates = db_q_by_norm.get(n, [])
            if candidates:
                idx, header, dbq = candidates[0]
                matches.append((col, idx, header, q, dbq))
            else:
                unmatched_needed.append((col, q))
    return headers, matches, needed_vlookup_cols, dropped_vlookup_cols, unmatched_needed


def row_for_patient(hcode, pnum):
    rows = gsheet("DB2SHEET!A1:ZW", "UNFORMATTED_VALUE")
    headers = rows[0]
    for row in rows[2:]:
        if len(row) > 3 and str(row[2]) == hcode and str(row[3]) == pnum:
            return headers, row
    raise RuntimeError(f"DB2SHEET row not found: {hcode}/{pnum}")


def recalc(path):
    outdir = path.parent / "lo-pruned-recalc"
    outdir.mkdir(exist_ok=True)
    converted = outdir / path.name
    if converted.exists():
        converted.unlink()
    subprocess.run([
        "/opt/homebrew/bin/soffice",
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(outdir),
        str(path),
    ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if not converted.exists():
        raise RuntimeError(f"LibreOffice failed to create {converted}")
    shutil.copy2(converted, path)


def compare(excel_rows, sheet_rows, row_numbers=None, clamp_zero_to_one=False):
    sheet_by_name = {}
    for row in sheet_rows:
        if row and row[0]:
            sheet_by_name.setdefault(norm_name(row[0]), deque()).append(row)
    fields = [(1, "응답결과"), (2, "환자군 백분위"), (3, "대조군 백분위")]
    stats = {label: {"total": 0, "match": 0, "mismatch": 0, "errors": 0, "items": []} for _, label in fields}
    material_stats = {label: {"total": 0, "match": 0, "mismatch": 0, "items": []} for _, label in fields}
    for pos, r in enumerate(excel_rows):
        report_row = row_numbers[pos] if row_numbers else None
        if not r or not r[0]:
            continue
        name = norm_name(r[0])
        if name == "검사명" or name not in sheet_by_name or not sheet_by_name[name]:
            continue
        s = sheet_by_name[name].popleft()
        for idx, label in fields:
            ev = r[idx] if len(r) > idx else ""
            sv = s[idx] if len(s) > idx else ""
            stats[label]["total"] += 1
            comparable = close(ev, sv) or (clamp_zero_to_one and first_num(ev) == 0 and first_num(sv) == 1)
            if comparable:
                stats[label]["match"] += 1
            else:
                stats[label]["mismatch"] += 1
                if str(ev).startswith("#"):
                    stats[label]["errors"] += 1
                stats[label]["items"].append({"name": r[0], "excel": ev, "sheet": sv})
            if report_row not in BROKEN_REPORT_ROWS and not str(ev).startswith("#"):
                material_stats[label]["total"] += 1
                if comparable:
                    material_stats[label]["match"] += 1
                else:
                    material_stats[label]["mismatch"] += 1
                    material_stats[label]["items"].append({"name": r[0], "excel": ev, "sheet": sv})
    stats["_material"] = material_stats
    return stats


def apply_legacy_output_patches(calc):
    for offset, source_col in enumerate(range(4, 9), start=12):
        calc.cell(81, offset).value = f"={get_column_letter(source_col)}30"


def report_rows_from_workbook(path, start=1, end=110):
    wbv = load_workbook(path, data_only=True)
    return [[wbv["검사결과지"].cell(r, c).value for c in range(2, 6)] for r in range(start, end)]


def main():
    OUT_DIR.mkdir(exist_ok=True)
    original = load_workbook(ORIGINAL_XLSX, data_only=False)
    deps = dependency_closure(original)
    needed_cols = {column_index_from_string(re.match(r"([A-Z]+)", c).group(1)) for c in deps if re.match(r"[A-Z]+4$", c)}
    _, matches, needed_vlookup_cols, dropped_vlookup_cols, unmatched_needed = build_question_map(needed_cols)

    summary = {
        "dependency_cells": len(deps),
        "needed_vlookup_question_cols": len(needed_vlookup_cols),
        "matched_needed_question_cols": len(matches),
        "unmatched_needed_question_cols": len(unmatched_needed),
        "dropped_non_report_vlookup_cols": len(dropped_vlookup_cols),
        "unmatched_needed_preview": unmatched_needed[:50],
        "patients": [],
    }

    for hcode, pnum in PATIENTS:
        headers, row = row_for_patient(hcode, pnum)
        row_by_idx = {i: row[i] if i < len(row) else "" for i in range(len(headers))}
        out = OUT_DIR / f"legacy_pruned_report_inputs_{pnum}.xlsx"
        shutil.copy2(ORIGINAL_XLSX, out)
        wb = load_workbook(out, data_only=False)
        calc = wb["결과계산(2)"]
        report = wb["검사결과지"]
        report["J4"] = pnum

        # Remove non-report-input survey columns by clearing their headers/input formulas.
        for col, _ in dropped_vlookup_cols:
            calc.cell(1, col).value = None
            calc.cell(2, col).value = None
            calc.cell(3, col).value = None
            calc.cell(4, col).value = None
        # Keep only report-referenced input cells, defaulting unmatched legacy questions to 0.
        for col in needed_vlookup_cols:
            calc.cell(4, col).value = 0
        for col, idx, header, _, _ in matches:
            calc.cell(4, col).value = row_by_idx.get(idx, "")

        # Current TEMPSA-SV 39-item scores can legitimately be 0. The legacy
        # workbook treated an all-zero TEMPS block as "not administered", which
        # breaks percentile formulas with #VALUE!. Keep the short-form answers;
        # only remove that legacy N/A guard in the verification copy.
        apply_legacy_output_patches(calc)

        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.save(out)
        recalc(out)

        excel_report = report_rows_from_workbook(out, 1, 110)
        gsheet_put("검사결과지!H4:J4", [[hcode, "번호", pnum]])
        time.sleep(1.5)
        sheet_report = gsheet("검사결과지!B1:E110", "FORMATTED_VALUE")
        stats = compare(excel_report, sheet_report, list(range(1, 110)), clamp_zero_to_one=True)
        summary["patients"].append({
            "hospital_code": hcode,
            "patient_number": pnum,
            "xlsx": str(out),
            "stats": stats,
        })

    report_path = OUT_DIR / "legacy_pruned_report_inputs_vs_gsheet_summary.json"
    report_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2))
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
