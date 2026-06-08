import json
import re
import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook

from legacy_excel_crosswalk_verify import (
    BROKEN_REPORT_ROWS,
    ORIGINAL_XLSX,
    OUT_DIR,
    TESTDATA_XLSX,
    apply_legacy_output_patches,
    close,
    first_num,
    recalc,
    report_rows_from_workbook,
)

PDF_DIR = Path("/Users/wm/Library/CloudStorage/OneDrive-개인/9. Temporal Work/2026 설문웹앱개밸")
PDFS = [
    PDF_DIR / "기분장애클리닉 기본 설문지 (우울, 불안, 정서기질)(응답) - 검사결과지.pdf",
    PDF_DIR / "기분장애클리닉 기본 설문지 (우울, 불안, 정서기질)(응답) - 검사결과지 (1).pdf",
    PDF_DIR / "기분장애클리닉 기본 설문지 (우울, 불안, 정서기질)(응답) - 검사결과지 (2).pdf",
]


def norm(value):
    return re.sub(r"\s+", "", str(value or ""))


def pdf_text(path):
    return subprocess.check_output(["/opt/homebrew/bin/pdftotext", "-layout", str(path), "-"], text=True)


def pdf_patient_number(text):
    m = re.search(r"등록번호:\s*([0-9]+)", text)
    return m.group(1) if m else ""


def line_numbers(line):
    return [float(x) for x in re.findall(r"-?\d+(?:\.\d+)?", line)]


def make_excel_from_test_row(row_num):
    OUT_DIR.mkdir(exist_ok=True)
    out = OUT_DIR / f"legacy_testdata_row_{row_num}.xlsx"
    shutil.copy2(ORIGINAL_XLSX, out)
    src = load_workbook(TESTDATA_XLSX, data_only=True, read_only=True)["Sheet1"]
    wb = load_workbook(out, data_only=False)
    calc = wb["결과계산(2)"]
    report = wb["검사결과지"]
    patient_number = src.cell(row_num, 2).value
    report["J4"] = patient_number
    for col in range(3, calc.max_column + 1):
        source_col = calc.cell(2, col).value
        try:
            source_col = int(source_col)
        except (TypeError, ValueError):
            continue
        if 1 <= source_col <= src.max_column:
            calc.cell(4, col).value = src.cell(row_num, source_col).value
    apply_legacy_output_patches(calc)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(out)
    recalc(out)
    return out, str(patient_number)


def extract_pdf_rows(text, expected_rows):
    lines = [line.rstrip() for line in text.splitlines()]
    out = {}
    cursor = 0
    for report_row, row in expected_rows:
        if report_row in BROKEN_REPORT_ROWS or not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if not name or name == "검사명":
            continue
        key = (report_row, norm(name))
        target = norm(name)
        for idx in range(cursor, len(lines)):
            line = lines[idx]
            if norm(line).startswith(target):
                out[key] = line
                cursor = idx + 1
                break
    return out


def compare_pdf_to_excel(pdf_path, xlsx_path):
    text = pdf_text(pdf_path)
    expected_rows = list(zip(range(1, 110), report_rows_from_workbook(xlsx_path, 1, 110)))
    pdf_rows = extract_pdf_rows(text, expected_rows)
    stats = {"total": 0, "match": 0, "mismatch": 0, "missing": 0, "items": []}
    for report_row, row in expected_rows:
        if report_row in BROKEN_REPORT_ROWS or not row or not row[0] or row[0] == "검사명":
            continue
        expected = row[1:4]
        if all(first_num(v) is None for v in expected):
            continue
        key = (report_row, norm(row[0]))
        line = pdf_rows.get(key)
        stats["total"] += 1
        if not line:
            stats["missing"] += 1
            stats["items"].append({"row": report_row, "name": row[0], "issue": "missing"})
            continue
        nums = line_numbers(line)
        expected_nums = [first_num(v) for v in expected if first_num(v) is not None]
        # The line can include item numbers from descriptive text after the table
        # columns, so compare the first N numeric fields in table order.
        got_nums = nums[: len(expected_nums)]
        ok = len(got_nums) == len(expected_nums) and all(abs(a - b) <= 0.51 for a, b in zip(got_nums, expected_nums))
        if ok:
            stats["match"] += 1
        else:
            stats["mismatch"] += 1
            stats["items"].append({
                "row": report_row,
                "name": row[0],
                "excel": expected,
                "pdf_line": line.strip(),
                "pdf_nums": got_nums,
                "expected_nums": expected_nums,
            })
    return {
        "pdf": str(pdf_path),
        "pdf_patient_number": pdf_patient_number(text),
        "xlsx": str(xlsx_path),
        "stats": stats,
    }


def main():
    generated = []
    for row_num in (4, 5, 6):
        xlsx, patient_number = make_excel_from_test_row(row_num)
        generated.append({"row_num": row_num, "patient_number": patient_number, "xlsx": xlsx})

    by_patient = {item["patient_number"]: item for item in generated}
    results = []
    for pdf in PDFS:
        text = pdf_text(pdf)
        pnum = pdf_patient_number(text)
        item = by_patient.get(pnum)
        if not item:
            results.append({"pdf": str(pdf), "pdf_patient_number": pnum, "error": "matching testdata row not found"})
            continue
        results.append(compare_pdf_to_excel(pdf, item["xlsx"]))

    out = {"generated": [{**item, "xlsx": str(item["xlsx"])} for item in generated], "results": results}
    path = OUT_DIR / "legacy_pdf_cases_vs_excel_summary.json"
    path.write_text(json.dumps(out, ensure_ascii=False, indent=2))
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
